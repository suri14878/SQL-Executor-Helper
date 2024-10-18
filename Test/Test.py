import itertools
import unittest
import os,sys, shutil
import csv
import configparser 
from openpyxl import load_workbook

# Get the parent directory
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, parent_dir)
import Logger
from Executor.SQLExecutor import SQLExecutor, OracleConnection, PostgresConnection
from Executor.enums.file_types import FileType

class TestSQLExecutorIntegration(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        """Common Setup for initializing and testing databases."""

        # Load configuration settings
        config = cls.load_config()

        # Create Logger
        logging = Logger.create_root()
        cls.logger = logging.getLogger('Tester')
        cls.logger.info("Setting up the databases and test environment...")

        # Mocked Data
        cls.testActors_data =  [{'PK_ID': 1, 'NAME': 'Actor 1', 'SEX': 'Male', 'BIO': 'Bio of Actor 1'},
                            {'PK_ID': 2, 'NAME': 'Actor 2', 'SEX': 'Male', 'BIO': 'Bio of Actor 2'},
                            {'PK_ID': 3, 'NAME': 'Actor 3', 'SEX': 'Male', 'BIO': 'Bio of Actor 3'},
                            {'PK_ID': 4, 'NAME': 'Actor 4', 'SEX': 'Male', 'BIO': 'Bio of Actor 4'},
                            {'PK_ID': 5, 'NAME': 'Actor 5', 'SEX': 'Male', 'BIO': 'Bio of Actor 5'},
                            {'PK_ID': 6, 'NAME': 'Actor 6', 'SEX': 'Male', 'BIO': 'Bio of Actor 6'},
                            {'PK_ID': 7, 'NAME': 'Actor 7', 'SEX': 'Male', 'BIO': 'Bio of Actor 7'},
                            {'PK_ID': 8, 'NAME': 'Actor 8', 'SEX': 'Male', 'BIO': 'Bio of Actor 8'},
                            {'PK_ID': 9, 'NAME': 'Actor 9', 'SEX': 'Male', 'BIO': 'Bio of Actor 9'},
                            {'PK_ID': 10, 'NAME': 'Actor 10', 'SEX': 'Male', 'BIO': 'Bio of Actor 10'}]

        # Can comment any database if want to test only the other.
        cls.databases = {}
        
        if config.getboolean('TestSettings', 'TestOracle'):
            cls.databases['oracle'] = SQLExecutor(OracleConnection(), config_file='./Configs/Database_Config.ini', environment='test')
        
        if config.getboolean('TestSettings', 'TestPostgres'):
            cls.databases['postgres'] = SQLExecutor(PostgresConnection(), config_file='./Configs/Database_Config.ini', environment='test')

        # Connect to the databases
        for db_type, db in cls.databases.items():
            try:
                cls.logger.info(f"Connected to {db_type} database successfully.")
                if db_type == 'oracle':
                    cls.setup_oracle(db)
                elif db_type == 'postgres':
                    cls.setup_postgres(db)
            except Exception as e:
                cls.logger.error(f"Failed to connect to {db_type} database: {e}")
        
        os.mkdir('./Test/TestFiles')
        os.mkdir('./Test/TestFiles/oracle')
        os.mkdir('./Test/TestFiles/postgres')
        os.mkdir('./Test/SQL Files')
        os.mkdir('./Test/TestScripts')

        # Create necessary SQL files for testing
        cls.logger.info("Creating test SQL files...")

        try:
            with open('./Test/SQL Files/singleQuery.sql', 'w') as file:
                    file.write('SELECT * FROM TestActors')
            with open('./Test/SQL Files/multipleQueries.sql', 'w') as file:
                file.write('''SELECT * FROM TestActors;
                        /* 
                            This query is just for testing purposes
                        */
                        /* PAGINATE SIZE 3 */
                        SELECT * FROM TestActors;
                        /* This is a test query */
                        SELECT * FROM TestActors
                        WHERE "PK_ID" = 5;
                            ''')
            with open('./Test/SQL Files/multipleQueries_with_params.sql', 'w') as file:
                file.write('''SELECT * FROM TestActors;
                        /* 
                            This query is just for testing purposes
                        */
                        /* PAGINATE SIZE 3 */
                        /* ROW LIMIT 6 */
                        SELECT * FROM TestActors;
                        /* This is a test query */
                        SELECT * FROM TestActors
                        WHERE "PK_ID" IN (3,4,5);
                            ''')
            with open('./Test/TestScripts/Adarsh.sql', 'w') as file:
                file.write('''SELECT * FROM TestActors;
                        /* 
                            This query is just for testing purposes
                        */
                        /* PAGINATE SIZE 2 */
                        SELECT * FROM TestActors;
                            ''')
            with open('./Test/TestScripts/Blake.sql', 'w') as file:
                file.write('''/* This is a test query */
                        SELECT * FROM TestActors
                        WHERE "PK_ID" IN (3,4,5);
                        SELECT * FROM TestActors;
                            ''')
            cls.logger.info("Test SQL files created successfully.")
        except Exception as e:
            cls.logger.error(f"Error creating test SQL files: {e}")

    @classmethod
    def setup_oracle(cls, db):
        """Setup Oracle-specific tables and data."""
        cls.logger.info("Setting up Oracle database...")
        try:
            with db.transaction():
                db.execute_query("""
                BEGIN
                EXECUTE IMMEDIATE 'DROP TABLE TestActors CASCADE CONSTRAINTS';
                EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -942 THEN
                        RAISE;
                    END IF;
                END;
                """)
                db.execute_query("""
                    CREATE TABLE TestActors (
                        PK_ID INTEGER PRIMARY KEY,
                        NAME VARCHAR(100),
                        SEX VARCHAR(10),
                        BIO VARCHAR(1000)
                    )
                """)
                for i in range(1, 11):
                    db.execute_query(f"""
                        INSERT INTO TestActors (PK_ID, NAME, SEX, BIO)
                        VALUES ({i}, 'Actor {i}', 'Male', 'Bio of Actor {i}')
                    """)
            cls.logger.info("Oracle database setup completed.")
        except Exception as e:
            cls.logger.error(f"Error setting up Oracle database: {e}")

    @classmethod
    def setup_postgres(cls, db):
        """Setup Postgres-specific tables and data."""
        cls.logger.info("Setting up Postgres database...")
        try:
            with db.transaction():
                db.execute_query("DROP TABLE IF EXISTS TestActors CASCADE;")
                db.execute_query("""
                    CREATE TABLE TestActors (
                        "PK_ID" SERIAL PRIMARY KEY,
                        "NAME" VARCHAR(100),
                        "SEX" VARCHAR(10),
                        "BIO" TEXT
                    )
                """)

                for i in range(1, 11):
                    db.execute_query(f"""
                        INSERT INTO TestActors ("NAME", "SEX", "BIO")
                        VALUES ('Actor {i}', 'Male', 'Bio of Actor {i}')
                    """)
                cls.logger.info("Postgres database setup completed.")
        except Exception as e:
            cls.logger.error(f"Error setting up Postgres database: {e}")


    def test_fetchAll_save_to_files(self):
        """Test Database features: save to CSV, TXT, Excel."""
        output_dir = './Test/TestFiles'
        for db_type, db in self.databases.items():
            with self.subTest(db=db_type):
                self.logger.info(f"Running fetchAll test for {db_type} database.")
                try:
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchAll_test'),result_file_type=FileType.CSV)
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchAll_test'),result_file_type=FileType.TXT)
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchAll_test'),result_file_type=FileType.EXCEL)
                    self.verify_files(self.testActors_data, db_type, 'fetchAll')
                except Exception as e:
                    self.logger.error(f"Error in fetchAll test for {db_type}: {e}")
                    raise

    def test_fetchOne_save_to_files(self):
        """Test fetching one record from both databases and saving to CSV, TXT, Excel."""
        output_dir = './Test/TestFiles'
        for db_type, db in self.databases.items():
            with self.subTest(db=db_type):
                self.logger.info(f"Running fetchOne test for {db_type} database.")
                try:
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchOne_test'),result_file_type=FileType.CSV, batch_size=1)
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchOne_test'),result_file_type=FileType.TXT, batch_size=1)
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchOne_test'),result_file_type=FileType.EXCEL, batch_size=1)
                    self.verify_files(self.testActors_data, db_type, 'fetchOne')
                except Exception as e:
                    self.logger.error(f"Error in fetchOne test for {db_type}: {e}")
                    raise

    def test_fetchMany_save_to_files(self):
        """Test fetching multiple records from both databases and saving to CSV, TXT, Excel."""
        output_dir = './Test/TestFiles'
        for db_type, db in self.databases.items():
            with self.subTest(db=db_type):
                self.logger.info(f"Running fetchMany test for {db_type} database.")
                try:
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchMany_test'),result_file_type=FileType.CSV, batch_size=3)
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchMany_test'),result_file_type=FileType.TXT, batch_size=3)
                    db.execute_file_and_save(file_name='./Test/SQL Files/singleQuery.sql',result_file_path=os.path.join(output_dir, f'{db_type}_fetchMany_test'),result_file_type=FileType.EXCEL, batch_size=3)
                    self.verify_files(self.testActors_data, db_type, 'fetchMany')
                except Exception as e:
                    self.logger.error(f"Error in fetchMany test for {db_type}: {e}")
                    raise

    def test_multiquery_to_file(self):
        """Test fetching multiple queries from single file from both databases and saving to CSV, TXT, Excel."""
        list_data =[self.testActors_data, self.testActors_data, [self.testActors_data[4]]]
        for db_type, db in self.databases.items():
            output_path = f'./Test/TestFiles/{db_type}_test'
            with self.subTest(db=db_type):
                self.logger.info(f"Running multi-query test for {db_type} database.")
                try:
                    db.execute_file_and_save(file_name='./Test/SQL Files/multipleQueries.sql', result_file_path=output_path, result_file_type=FileType.CSV)
                    db.execute_file_and_save(file_name='./Test/SQL Files/multipleQueries.sql', result_file_path=output_path, result_file_type=FileType.TXT)
                    db.execute_file_and_save(file_name='./Test/SQL Files/multipleQueries.sql', result_file_path=output_path, result_file_type=FileType.EXCEL)
                    self.verify_multiQuery_files(db_type, list_data, 'test')
                except Exception as e:
                    self.logger.error(f"Error in multi-query test for {db_type}: {e}")
                    raise

    def test_multiquery_with_params_to_file(self):
        """Test fetching multiple queries from single file from both databases and saving to CSV, TXT, Excel."""
        list_data =[self.testActors_data[:5], self.testActors_data[:6], self.testActors_data[2:5]]
        for db_type, db in self.databases.items():
            output_path = f'./Test/TestFiles/{db_type}_test_params'
            with self.subTest(db=db_type):
                self.logger.info(f"Running multi-query test for {db_type} database.")
                try:
                    db.execute_file_and_save(file_name='./Test/SQL Files/multipleQueries_with_params.sql', result_file_path=output_path, result_file_type=FileType.CSV, row_limit=5, batch_size=2)
                    db.execute_file_and_save(file_name='./Test/SQL Files/multipleQueries_with_params.sql', result_file_path=output_path, result_file_type=FileType.TXT, row_limit=5, batch_size=2)
                    db.execute_file_and_save(file_name='./Test/SQL Files/multipleQueries_with_params.sql', result_file_path=output_path, result_file_type=FileType.EXCEL, row_limit=5, batch_size=2)
                    self.verify_multiQuery_files(db_type, list_data, 'test_params')
                except Exception as e:
                    self.logger.error(f"Error in multi-query test for {db_type}: {e}")
                    raise

    def test_folder_to_file_and_verify_files(self):
        """Test fetching multiple files from folder from both databases and saving to CSV, TXT, Excel."""
        folder_path = './Test/TestScripts'
        adarsh_file_data = [self.testActors_data, self.testActors_data]
        blake_file_data = [self.testActors_data[2:5], self.testActors_data]
        final_data = {'Adarsh':adarsh_file_data, 'Blake': blake_file_data}
        for db_type, db in self.databases.items():
            result_save_path = f'./Test/TestFiles/{db_type}/'
            with self.subTest(db=db_type):
                self.logger.info(f"Running folder query test for {db_type} database.")
                try:
                    # Check if the provided path is a valid directory
                    if not os.path.isdir(folder_path):
                        self.logger.warning(f"Directory {folder_path} does not exist.")
                        return
                    
                    db.execute_folder_and_save(folder_path, result_save_path , result_file_type=FileType.CSV)
                    db.execute_folder_and_save(folder_path, result_save_path, result_file_type=FileType.TXT)
                    db.execute_folder_and_save(folder_path, result_save_path, result_file_type=FileType.EXCEL)
                    self.verify_folder_files(db_type, final_data)
                except Exception as e:
                    self.logger.error(f"Error in folder query test for {db_type}: {e}")
                    raise

    def test_transaction_rollback_on_error(self):
        """Test that a transaction rolls back on error, specifically on a division by zero error."""
        for db_type, db in self.databases.items():
            with self.subTest(db=db_type):
                self.logger.info(f"Running transaction rollback test for {db_type} database.")
                
                try:
                    # Start a transaction
                    with db.transaction():
                        if db_type == 'postgres':
                            db.execute_query("INSERT INTO TestActors (\"PK_ID\", \"NAME\", \"SEX\", \"BIO\") VALUES (900, 'Actor 900', 'Male', 'Should be rolled back')")
                            db.execute_query("SELECT 1/0 AS db_exception")
                        elif db_type == 'oracle':
                            db.execute_query("INSERT INTO TestActors (PK_ID, NAME, SEX, BIO) VALUES (900, 'Actor 900', 'Male', 'Should be rolled back')")
                            db.execute_query("SELECT 1/0 AS db_exception FROM dual")

                    # If no exception was raised, fail the test
                    self.fail(f"Expected division by zero exception in {db_type} did not occur.")

                except Exception as e:
                    # Expected error occurred, check rollback
                    self.logger.info(f"Division by zero error occurred as expected in {db_type}: {e}")

                    # Now check that the insert was rolled back
                    batches = db.get_batches_by_query("SELECT * FROM TestActors WHERE \"PK_ID\" = 900", page_size = 1)
                    rows = [row for row in itertools.chain.from_iterable(batches)]
                    self.assertEqual(len(rows), 0, f"Transaction did not roll back in {db_type}; found row with PK_ID 900.")

                else:
                    # If no exception was caught, this means the transaction did not rollback as expected
                    self.fail(f"Transaction did not roll back in {db_type} as expected on division by zero.")

    # def test_retry_on_insert_after_connection_loss(self):
    #     """Test retry mechanism when the database connection is lost between the process."""
    #     for db_type, db in self.databases.items():
    #         with self.subTest(db=db_type):
    #             self.logger.info(f"Running connection terminate test for {db_type} database.")

    #             try:
    #                 # Start a transaction
    #                 with db.transaction():
    #                     if db_type == 'postgres':
    #                         db.execute_query("INSERT INTO TestActors (\"PK_ID\", \"NAME\", \"SEX\", \"BIO\") VALUES (902, 'Actor 902', 'Male', 'Should be rolled back')")
    #                         db.execute_query("""
    #                                             SELECT pg_terminate_backend(pid)
    #                                             FROM pg_stat_activity
    #                                             WHERE pid <> pg_backend_pid()
    #                                             AND datname = 'test_database_name'; 
    #                                         """)
    #                     elif db_type == 'oracle':
    #                         db.execute_query("INSERT INTO TestActors (PK_ID, NAME, SEX, BIO) VALUES (902, 'Actor 902', 'Male', 'Should be rolled back')")
    #                         db.execute_query("SELECT 1/0 AS db_exception FROM dual")

    #                 # If no exception was raised, fail the test
    #                 self.fail(f"Expected division by zero exception in {db_type} did not occur.")

    #             except Exception as e:
    #                 # Expected error occurred, check rollback
    #                 self.logger.info(f"Division by zero error occurred as expected in {db_type}: {e}")

    #                 # Now check that the insert was rolled back
    #                 batches = db.get_batches_by_query("SELECT * FROM TestActors WHERE \"PK_ID\" = 902", page_size = 1)
    #                 rows = [row for row in itertools.chain.from_iterable(batches)]
    #                 self.assertEqual(len(rows), 0, f"Transaction did not roll back in {db_type}; found row with PK_ID 902.")

    #             else:
    #                 # If no exception was caught, this means the transaction did not rollback as expected
    #                 self.fail(f"Transaction did not roll back in {db_type} as expected on division by zero.")

    def verify_multiQuery_files(self, db_type, list_data, file_name):
        """Helper method to save multiquery data and verify CSV, TXT, Excel files."""
        output_dir = './Test/TestFiles/'
        
        for i, data in enumerate(list_data):
            # Check if files exist
            self.assertTrue(os.path.exists(os.path.join(output_dir, f'{db_type}_{file_name}_{i+1}.csv')))
            self.assertTrue(os.path.exists(os.path.join(output_dir, f'{db_type}_{file_name}_{i+1}.txt')))
            self.assertTrue(os.path.exists(os.path.join(output_dir, f'{db_type}_{file_name}_{i+1}.xlsx')))

            # Verify file contents
            self.verify_csv_content(os.path.join(output_dir, f'{db_type}_{file_name}_{i+1}.csv'), data)
            self.verify_txt_content(os.path.join(output_dir, f'{db_type}_{file_name}_{i+1}.txt'), data)
            self.verify_excel_content(os.path.join(output_dir, f'{db_type}_{file_name}_{i+1}.xlsx'), data)

    def verify_folder_files(self, db_type, final_data):
        """Helper method to save multiquery data and verify CSV, TXT, Excel files."""
        output_dir = f'./Test/TestFiles/{db_type}/'

        for file_name, value in final_data.items():
            for i, data in enumerate(value):
                # # Check if files exist
                self.assertTrue(os.path.exists(os.path.join(output_dir, f'{file_name}_{i+1}.csv')))
                self.assertTrue(os.path.exists(os.path.join(output_dir, f'{file_name}_{i+1}.txt')))
                self.assertTrue(os.path.exists(os.path.join(output_dir, f'{file_name}_{i+1}.xlsx')))

                # # Verify file contents
                self.verify_csv_content(os.path.join(output_dir, f'{file_name}_{i+1}.csv'), data)
                self.verify_txt_content(os.path.join(output_dir, f'{file_name}_{i+1}.txt'), data)
                self.verify_excel_content(os.path.join(output_dir, f'{file_name}_{i+1}.xlsx'), data)

    def verify_files(self,data, db_type, fetch_type):
        """Helper method to save data and verify CSV, TXT, Excel files."""
        output_dir = './Test/TestFiles'

        # Check if files exist
        self.assertTrue(os.path.exists(os.path.join(output_dir, f'{db_type}_{fetch_type}_test_1.csv')))
        self.assertTrue(os.path.exists(os.path.join(output_dir, f'{db_type}_{fetch_type}_test_1.txt')))
        self.assertTrue(os.path.exists(os.path.join(output_dir, f'{db_type}_{fetch_type}_test_1.xlsx')))

        # Verify file contents
        self.verify_csv_content(os.path.join(output_dir, f'{db_type}_{fetch_type}_test_1.csv'), data)
        self.verify_txt_content(os.path.join(output_dir, f'{db_type}_{fetch_type}_test_1.txt'), data)
        self.verify_excel_content(os.path.join(output_dir, f'{db_type}_{fetch_type}_test_1.xlsx'), data)

    # Verification methods...
    def verify_csv_content(self, file_path, expected_data):
        """Verify the content of the CSV file."""
        with open(file_path, mode='r') as file:
            reader = csv.DictReader(file)
            rows = list(reader)
            self.assertEqual(len(rows), len(expected_data))
            for i, row in enumerate(rows):
                for key, value in row.items():
                    self.assertEqual(str(value), str(expected_data[i][key]) if  str(expected_data[i][key]) != 'None' else '')

    def verify_txt_content(self, file_path, expected_data):
        """Verify the content of the TXT file."""
        with open(file_path, mode='r') as file:
            lines = file.readlines()
            self.assertEqual(len(lines) - 1, len(expected_data))  # -1 for the header
            headers = lines[0].strip().split('\t')
            for i, line in enumerate(lines[1:]):
                values = line.strip().split('\t')
                row = dict(zip(headers, values))
                for key in headers:
                    self.assertEqual(str(row[key]), str(expected_data[i][key]))

    def verify_excel_content(self, file_path, expected_data):
        """Verify the content of the Excel file."""
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        for i, row in enumerate(sheet.iter_rows(values_only=True, min_row=2)):
            row_data = dict(zip(headers, row))
            for key in headers:
                self.assertEqual(str(row_data[key]), str(expected_data[i][key]) if  str(expected_data[i][key]) != 'None' else '')

    @classmethod
    def load_config(cls):
        """Load the configuration settings from TestConfig.ini."""
        config = configparser.ConfigParser()
        config.read('./Configs/Test_Config.ini')
        return config
    
    # Deleting all the testfiles and close connections
    @classmethod
    def tearDownClass(cls):
        """Tear down the test environment: drop the test tables and clean up."""

        config = cls.load_config()

        cls.logger.info("Tearing down test environment...")
        for db_type, db in cls.databases.items():
            try:
                with db.transaction():
                    if db_type == 'oracle':
                        db.execute_query("DROP TABLE TestActors CASCADE CONSTRAINTS")
                    elif db_type == 'postgres':
                        db.execute_query("DROP TABLE IF EXISTS TestActors CASCADE")
                    cls.logger.info(f"{db_type} test table dropped.")
            except Exception as e:
                cls.logger.error(f"Error dropping {db_type} table: {e}")

        # Decide whether to clean up files based on config
        if config.getboolean('TestSettings', 'CleanUpFiles'):
            output_dir = './Test/TestFiles/'
            sqlFiles_dir = './Test/SQL Files'
            testScripts_dir = './Test/TestScripts'
            try:
                if os.path.exists(output_dir):
                    shutil.rmtree(output_dir)
                if os.path.exists(sqlFiles_dir):
                    shutil.rmtree(sqlFiles_dir)
                if os.path.exists(testScripts_dir):
                    shutil.rmtree(testScripts_dir)
                cls.logger.info("Test directories and files cleaned up.")
            except Exception as e:
                cls.logger.error(f"Error during test cleanup: {e}")

if __name__ == '__main__':
    unittest.main()
