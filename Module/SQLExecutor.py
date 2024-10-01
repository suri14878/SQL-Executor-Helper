import os
import oracledb
import psycopg
import configparser
import csv
import re
import uuid
from openpyxl import Workbook, load_workbook
from Module.Helpers import Logger
from Module.enums.file_types import FileType


logging = Logger.create_logger()

class UniqueDictRowFactory:
    def __init__(self, cursor: psycopg.Cursor):
        # Check if the query returns a result set (i.e., cursor.description is not None)
        if cursor.description is None:
            self.fields = None  # No fields to process if there's no result set
        else:
            # Extract column names from cursor.description
            self.fields = []
            field_count = {}
            for col in cursor.description:
                col_name = col.name
                if col_name in field_count:
                    # If the column name already exists, append a unique identifier
                    field_count[col_name] += 1
                    col_name = f"{col_name}_{field_count[col_name]}"
                else:
                    field_count[col_name] = 0
                self.fields.append(col_name)

    def __call__(self, values):
        # If there's no result set (fields are None), return None or an empty dict
        if self.fields is None:
            return {}
        # Otherwise, create a dictionary mapping the unique column names to values
        return dict(zip(self.fields, values))


# Abstract base class defining the general interface for all database connections
class GeneralConnection:
    def __init__(self):
        raise NotImplementedError("Subclass must implement this method")
    
    def connect(self, config_file, environment):
        raise NotImplementedError("Subclass must implement this method")

    def close(self):
        raise NotImplementedError("Subclass must implement this method")

    def get_cursor(self):
        raise NotImplementedError("Subclass must implement this method")
    
# Abstract base class defining the general interface for all database cursors
class GeneralCursor:
    def execute(self, query, params=None):
        raise NotImplementedError("Subclass must implement this method")
    
    def fetchall(self):
        raise NotImplementedError("Subclass must implement this method")
    
    def fetchmany(self, size):
        raise NotImplementedError("Subclass must implement this method")
    
    def fetchone(self):
        raise NotImplementedError("Subclass must implement this method")
    
    @property
    def description(self):
        """Expose the description attribute from the underlying cursor."""
        raise NotImplementedError("Subclass must implement this method")

# Concrete class for PostgreSQL database connection
class PostgresConnection(GeneralConnection):
    def __init__(self):
        self.logger = logging.getLogger("Postgres")
        self.__connection = None

    def connect(self, config_file, environment):
        """Establishes a connection to a PostgreSQL database using the provided config file and environment."""
        try:
            config = configparser.ConfigParser()
            config.read(config_file)
            env_section = f'{environment}_postgres'
            self.__connection = psycopg.connect(
                host=config[env_section]['host'],
                port=config[env_section]['port'],
                user=config[env_section]['user'],
                password=config[env_section]['password'],
                dbname=config[env_section]['dbname'],
                row_factory=UniqueDictRowFactory
            )
            self.logger.info(f"Connected to PostgreSQL database: {config[env_section]['dbname']}")

        except (psycopg.OperationalError, configparser.Error) as e:
            self.logger.error(f"Failed to connect to PostgreSQL: {str(e)}")
            raise

    def close(self):
        """Closes the PostgreSQL database connection."""
        if self.__connection:
            self.__connection.close()
            self.logger.info("PostgreSQL connection closed.")

    def get_cursor(self, is_client_cursor = False):
        """Returns a generalized cursor object based on param for PostgreSQL."""
        try:
            if is_client_cursor:
                # Always return a client cursor (client-side)
                return PostgresCursor(self.__connection.cursor())
            else:
                return PostgresCursor(self.__connection.cursor(name=f"server_side_cursor_{uuid.uuid4()}"))
        except Exception as e:
            self.logger.error(f"Failed to create PostgreSQL cursor: {str(e)}")
            raise

# Concrete class for PostgreSQL cursor
class PostgresCursor(GeneralCursor):
    def __init__(self, cursor):
        self.logger = logging.getLogger("Postgres")
        self.__cursor = cursor

    def __enter__(self):
        """Enters the context manager and returns the cursor."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Closes the cursor when exiting the context."""
        self.__cursor.close()

    def execute(self, query, params=None):
        """Executes a SQL query using the PostgreSQL cursor."""
        try:
            self.__cursor.execute(query, params)
            self.logger.info("Executed query on PostgreSQL.")
        except psycopg.Error as e:
            self.logger.error(f"Failed to execute query on PostgreSQL: {str(e)}")
            raise

    def fetchone(self):
        """Fetches one row from the query result."""
        return self.__cursor.fetchone()

    def fetchall(self):
        """Fetches all rows from the query result."""
        return self.__cursor.fetchall()

    def fetchmany(self, size):
        """Fetches a specific number of rows from the query result."""
        return self.__cursor.fetchmany(size)

    @property
    def description(self):
        """Expose the description attribute from the underlying cursor."""
        return self.__cursor.description

# Concrete class for Oracle database connection
class OracleConnection(GeneralConnection):
    def __init__(self):
        self.logger = logging.getLogger("Oracle")
        self.__connection = None

    def connect(self, config_file, environment):
        """Establishes a connection to an Oracle database using the provided config file and environment."""
        try:
            config = configparser.ConfigParser()
            config.read(config_file)
            env_section = f'{environment}_oracle'
            dsn = oracledb.makedsn(
                config[env_section]['host'],
                config[env_section]['port'],
                sid=config[env_section]['sid']
            )
            self.__connection = oracledb.connect(
                user=config[env_section]['user'],
                password=config[env_section]['password'],
                dsn=dsn
            )
            self.logger.info(f"Connected to Oracle database: {config[env_section]['sid']}")

        except (oracledb.DatabaseError, configparser.Error) as e:
            self.logger.error(f"Failed to connect to Oracle: {str(e)}")
            raise

    def close(self):
        """Closes the Oracle database connection."""
        if self.__connection:
            self.__connection.close()
            self.logger.info("Oracle connection closed.")

    def get_cursor(self):
        """Returns a generalized cursor object for Oracle."""
        try:
            return OracleCursor(self.__connection.cursor())
        except Exception as e:
            self.logger.error(f"Failed to create Oracle cursor: {str(e)}")
            raise

# Concrete class for Oracle cursor
class OracleCursor(GeneralCursor):
    def __init__(self, cursor):
        self.logger = logging.getLogger("Oracle")
        self.__cursor = cursor

    def __enter__(self):
        """Enters the context manager and returns the cursor."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Closes the cursor when exiting the context."""
        self.__cursor.close()

    def execute(self, query, params=None):
        """Executes a SQL query using the Oracle cursor."""
        try:
            self.__cursor.execute(query, params or {})
            self.logger.info("Executed query on Oracle.")

        except oracledb.DatabaseError as e:
            self.logger.error(f"Failed to execute query on Oracle: {str(e)}")
            raise

    def fetchone(self):
        """Fetches one row from the query result."""
        self.__apply_row_factory()
        return self.__cursor.fetchone()

    def fetchall(self):
        """Fetches all rows from the query result."""
        self.__apply_row_factory()
        return self.__cursor.fetchall()

    def fetchmany(self, size):
        """Fetches a specific number of rows from the query result."""
        self.__apply_row_factory()
        return self.__cursor.fetchmany(size)

    @property
    def description(self):
        """Expose the description attribute from the underlying cursor."""
        return self.__cursor.description
    
    def __apply_row_factory(self):
        """Applies the row factory to format rows as dictionaries with column names as keys, handling duplicate column names."""
        # Extract column names from cursor description
        columns = [col[0] for col in self.__cursor.description]

        # Create a dictionary to count occurrences of column names
        field_count = {}
        unique_columns = []

        # Handle duplicate column names by appending a suffix
        for col_name in columns:
            if col_name in field_count:
                # If the column name already exists, append a unique identifier
                field_count[col_name] += 1
                unique_col_name = f"{col_name}_{field_count[col_name]}"
            else:
                field_count[col_name] = 0
                unique_col_name = col_name
            unique_columns.append(unique_col_name)

        # Set rowfactory to map unique column names to values
        self.__cursor.rowfactory = lambda *args: dict(zip(unique_columns, args))


# SQLExecutor class manages the connection and execution of SQL queries
class SQLExecutor:
    def __init__(self, db_connection: GeneralConnection, config_file='config.ini', environment='test'):
        self.logger = logging.getLogger("Executor")
        self.__db_connection = db_connection
        self.__db_connection.connect(config_file, environment)

    def __del__(self):
        """Closes the connection when the object is destroyed."""
        self.__db_connection.close()

    def _get_cursor(self, is_client_cursor = None):
        """Returns a new cursor. Each method will call this to open a new cursor."""    
        if is_client_cursor:        
            return self.__db_connection.get_cursor(is_client_cursor)
        else:
            return self.__db_connection.get_cursor()



    def execute_file_and_save(self, file_name, result_file_path, result_file_type, batch_size=None, row_limit=None):
        """Executes SQL queries from a file."""
        try:
            with open(file_name, 'r') as file:
                queries = file.read().split(';')

            # This is to append count to the file name according to the number of queries a file have. Like (001, 002, 003, ...).
            preceding_zeros = len(str(len(queries)))

            # Loop through each query in the file
            for i,query in enumerate(queries):
                query = query.strip()
                if not query:  # Skip empty queries
                    continue

                # Check if the query includes pagination (e.g., /* PAGINATE SIZE <number> */)
                is_paginate, query_page_size = self.extract_pagination_info(query)
                is_rowlimit, query_row_limit = self.extract_row_limit_info(query)

                # Apply limit based of query comments or parameter
                apply_limit = query_row_limit if is_rowlimit else row_limit

                # Apply batch_size based of query comments or parameter
                apply_batch_size = query_page_size if is_paginate else batch_size

                if (apply_batch_size and apply_batch_size <= 0) or (apply_limit and apply_limit <= 0):
                    self.logger.warning(f"No data has been fetched since batch size or row limit is zero or lesser for query number {i+1} and filename: {file_name}")
                    continue

                # Handle queries with batch sizes (including 1)
                if apply_batch_size:
                    # Fetch rows in specified batch sizes
                    rows_fetched = 0
                    batches = self.get_batches_by_query(query, apply_batch_size)
                    batch_index = 0 
                    for batch in batches:
                        # This logic will check that batch_size should be within the row limits.
                        if apply_limit:
                            remaining_rows = apply_limit - rows_fetched
                            if remaining_rows < apply_batch_size:
                                batch = batch[:remaining_rows]
                        self.save_results(batch, f"{result_file_path}_{str(i + 1).zfill(preceding_zeros)}", result_file_type, is_append=True, include_header=True if batch_index==0 else False)
                        rows_fetched += len(batch)
                        batch_index += 1
                        if apply_limit and rows_fetched >= apply_limit:
                            break
                else:
                    with self._get_cursor() as cursor:
                        cursor.execute(query)
                        # Fetch all rows at once if no batch size specified
                        if apply_limit:
                            rows = cursor.fetchmany(apply_limit)
                        else:
                            rows = cursor.fetchall()
                        if result_file_type == FileType.CSV:
                            self.save_to_csv(rows,f"{result_file_path}_{str(i+1).zfill(preceding_zeros)}")    
                        elif result_file_type == FileType.TXT:
                            self.save_to_txt(rows,f"{result_file_path}_{str(i+1).zfill(preceding_zeros)}")    
                        elif result_file_type == FileType.EXCEL:
                            self.save_to_excel(rows,f"{result_file_path}_{str(i+1).zfill(preceding_zeros)}")
                        
        except FileNotFoundError as e:
            self.logger.error(f"SQL file not found: {file_name}")
            raise
        except Exception as e:
            self.logger.error(f"Failed to execute SQL file: {str(e)}")
            raise

    def execute_folder_and_save(self,folder_path, result_save_path, result_file_type, batch_size=None, row_limit=None):
        """Executes SQL files from a folder and saves the results."""
        
        # Check if the provided path is a valid directory
        if not os.path.isdir(folder_path):
            self.logger.error(f"The provided path '{folder_path}' is not a valid directory.")
            return

        # Loop through all files in the folder
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            if os.path.isfile(file_path):
                # Execute each file and save the results
                self.execute_file_and_save(file_path, f"{result_save_path}{os.path.splitext(file_name)[0]}", result_file_type, batch_size, row_limit)

    def extract_pagination_info(self, query):
        """Extracts the pagination information from multiline comments using the pattern '/* PAGINATE SIZE <number> */'."""
        match = re.search(r'/\*\s*PAGINATE\s+SIZE\s+(\d+)\s*\*/', query, flags=re.IGNORECASE)
        if match:
            return True, int(match.group(1))  # Return pagination flag and page size as an integer
        return False, None  # Return False if no pagination information is found
    
    def extract_row_limit_info(self, query):
        """Extracts the row limit information from multiline comments using the pattern '/* ROW LIMIT <number> */'."""
        match = re.search(r'/\*\s*ROW\s+LIMIT\s+(\d+)\s*\*/', query, flags=re.IGNORECASE)
        if match:
            return True, int(match.group(1))  # Return pagination flag and page size as an integer
        return False, None  # Return False if no pagination information is found

    def get_batches_by_query(self, query, page_size, params=None):
        """Gets query by batches."""
        with self._get_cursor() as cursor:
            cursor.execute(query, params)
            rows = cursor.fetchmany(page_size)
            while rows:
                yield rows
                rows = cursor.fetchmany(page_size)

    def get_queries_from_file(self, file_name):
        """Returns the list of queries from the SQL file."""
        with open(file_name, 'r') as file:
            queries = file.read().split(';')
        return [query.strip() for query in queries if query.strip()]
    
    def get_query_by_index(self, file_name, index):
        """Returns a query by its index in the file."""
        queries = self.get_queries_from_file(file_name)
        if index < len(queries):
            return queries[index]
        return None

    def save_results(self, data, result_file, result_file_type: FileType, is_append=False, include_header=True):
        """Appends these batches in specified file."""
        if result_file_type == FileType.CSV:
            self.save_to_csv(data,result_file, is_append, include_header)    
        elif result_file_type == FileType.TXT:
            self.save_to_txt(data,result_file, is_append, include_header)    
        elif result_file_type == FileType.EXCEL:
            self.save_to_excel(data,result_file, is_append, include_header)   

    def save_to_csv(self, data, file_name, is_append=False, include_header=True):
        """Saves data to a CSV file with column names."""
        try:
            if not data:  # Check if data is empty
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
        
            if isinstance(data, dict):  # If data is a single dictionary, wrap it in a list
                data = [data]

            # Check if the file name has a .csv extension, if not, add it
            if not file_name.lower().endswith('.csv'):
                file_name += '.csv'

            with open(file_name, 'a' if is_append else 'w', newline='') as file:
                writer = csv.DictWriter(file, fieldnames=data[0].keys())
                if include_header:
                    writer.writeheader()
                writer.writerows(data)
            self.logger.info(f"Data saved to {file_name} as CSV.")
        except Exception as e:
            self.logger.error(f"Failed to save data to CSV: {str(e)}")
            raise

    def save_to_txt(self, data, file_name, is_append=False, include_header=True):
        """Saves data to a TXT file with column names, using tab-separated values."""
        try:
            if not data:  # Check if data is empty
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
        
            if isinstance(data, dict):  # If data is a single dictionary, wrap it in a list
                data = [data]
            
            # Check if the file name has a .csv extension, if not, add it
            if not file_name.lower().endswith('.txt'):
                file_name += '.txt'

            with open(file_name, 'a' if is_append else 'w', newline='') as file:
                # Writing headers
                if include_header:
                    file.write('\t'.join(data[0].keys()) + '\n')
                # Writing data
                for row in data:
                    file.write('\t'.join(map(str, row.values())) + '\n')
            self.logger.info(f"Data saved to {file_name} as TXT.")
        except Exception as e:
            self.logger.error(f"Failed to save data to TXT: {str(e)}")
            raise

    def save_to_excel(self, data, file_name, is_append=False, include_header=True):
        """Saves data to an Excel file with column names."""
        try:
            if not data:  # Check if data is empty
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
        
            if not file_name.lower().endswith('.xlsx'):
                file_name += '.xlsx'

            if isinstance(data, dict):  # If data is a single dictionary, wrap it in a list
                data = [data]

            wb = None
            if is_append:
                try:
                    wb = load_workbook(file_name)
                except FileNotFoundError:
                    wb = Workbook()
            else:
                wb = Workbook()
            ws = wb.active
            # Writing headers
            if include_header:
                ws.append(list(data[0].keys()))
            # Writing data
            for row in data:
                ws.append(list(row.values()))
            wb.save(file_name)
            self.logger.info(f"Data saved to {file_name} as Excel.")
        except Exception as e:
            self.logger.error(f"Failed to save data to Excel: {str(e)}")
            raise()