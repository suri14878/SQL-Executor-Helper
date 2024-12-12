import os
import oracledb
import psycopg
import configparser
import csv
import re
import uuid
from openpyxl import Workbook, load_workbook
from Executor.enums.file_types import FileType
import logging
import time

# Retry decorator with exponential backoff
def retry(tries, delay=3, backoff=2, exceptions=(Exception,)):
    """
    Retry calling the decorated function using an exponential backoff.

    :param tries: Number of attempts to try (not retry) before giving up.
    :param delay: Initial delay between attempts in seconds.
    :param backoff: Multiplier applied to the delay after each retry.
    :param exceptions: Exceptions to catch and trigger retry.
    :raises: Last exception encountered.
    """
    if backoff <= 1:
        raise ValueError("backoff must be greater than 1")
    if tries < 0:
        raise ValueError("tries must be 0 or greater")
    if delay <= 0:
        raise ValueError("delay must be greater than 0")

    def deco_retry(f):
        def f_retry(self, *args, **kwargs):
            mtries, mdelay = tries, delay
            while mtries > 0:
                try:
                    return f(self, *args, **kwargs)
                except exceptions as e:
                    if self.is_terminated():
                        print(f"Retrying due to {e} (Remaining attempts: {mtries-1})...")
                        mtries -= 1
                        if mtries == 0:
                            raise  # If out of tries, raise the last exception
                        time.sleep(mdelay)
                        mdelay *= backoff
                    else:
                        raise
        return f_retry
    return deco_retry

# Retry decorator with exponential backoff
def retry_transaction(default_args, tries=3, delay=3, backoff=2, exceptions=(Exception,)):
    """
    Retry calling the decorated function using an exponential backoff and dynamic argument defaults.

    :param default_args: Dictionary mapping argument names to getter functions (lambdas) or direct values.
    :param tries: Number of attempts to try (not retry) before giving up.
    :param delay: Initial delay between attempts in seconds.
    :param backoff: Multiplier applied to the delay after each retry.
    :param exceptions: Exceptions to catch and trigger retry.
    :raises: Last exception encountered.
    """
    if backoff <= 1:
        raise ValueError("backoff must be greater than 1")
    if tries < 0:
        raise ValueError("tries must be 0 or greater")
    if delay <= 0:
        raise ValueError("delay must be greater than 0")

    def deco_retry(f):
        def f_retry(self, *args, **kwargs):
            # Create a new dictionary for arguments to avoid modifying original kwargs
            new_kwargs = kwargs.copy()

            # Populate new_kwargs dynamically using arg_getters
            for arg_name, getter in default_args.items():
                if callable(getter):
                    # If the getter is a lambda or callable function, invoke it
                    new_kwargs[arg_name] = getter(self)
                else:
                    # If the getter is not callable, assume it's a direct value
                    new_kwargs[arg_name] = getter

            mtries, mdelay = tries, delay
            while mtries > 0:
                try:
                    return f(self, *args, **kwargs)
                except exceptions as e:
                    if new_kwargs['db'].is_terminated():
                        print(f"Retrying due to Error: {e} (Remaining attempts: {mtries-1})...")
                        mtries -= 1
                        if mtries == 0:
                            raise  # If out of tries, raise the last exception
                        time.sleep(mdelay)
                        new_kwargs['db'].connect(new_kwargs['config_file'], new_kwargs['environment'])
                        mdelay *= backoff
                    else:
                        raise
        return f_retry
    return deco_retry

class UniqueDictRowFactory:
    def __init__(self, cursor: psycopg.Cursor):
        # Check if the query returns a result set (i.e., cursor.description is not None)
        if cursor.description is None:
            self.fields = None  # No fields to process if there's no result set
        else:
            # Extract column names from cursor.description
            self.fields = []
            field_count = {}
            for col in cursor.description:
                col_name = col.name
                if col_name in field_count:
                    # If the column name already exists, append a unique identifier
                    field_count[col_name] += 1
                    col_name = f"{col_name}_{field_count[col_name]}"
                else:
                    field_count[col_name] = 0
                self.fields.append(col_name)

    def __call__(self, values):
        # If there's no result set (fields are None), return None or an empty dict
        if self.fields is None:
            return {}
        # Otherwise, create a dictionary mapping the unique column names to values
        return dict(zip(self.fields, values))


# Abstract base class defining the general interface for all database connections
class GeneralConnection:
    def __init__(self):
        raise NotImplementedError("Subclass must implement this method")
    
    def connect(self, config_file, environment):
        raise NotImplementedError("Subclass must implement this method")

    def close(self):
        raise NotImplementedError("Subclass must implement this method")
    
    def commit(self):
        raise NotImplementedError("Subclass must implement this method")

    def get_cursor(self):
        raise NotImplementedError("Subclass must implement this method")
    
    def transaction(self):
        return Transaction(self)
    
# Abstract base class defining the general interface for all database cursors
class GeneralCursor:
    def execute(self, query, params=None):
        raise NotImplementedError("Subclass must implement this method")
    
    def fetchall(self):
        raise NotImplementedError("Subclass must implement this method")
    
    def fetchmany(self, size):
        raise NotImplementedError("Subclass must implement this method")
    
    def fetchone(self):
        raise NotImplementedError("Subclass must implement this method")
    
    @property
    def description(self):
        """Expose the description attribute from the underlying cursor."""
        raise NotImplementedError("Subclass must implement this method")

# Concrete class for PostgreSQL database connection
class PostgresConnection(GeneralConnection):
    def __init__(self):
        self.logger = logging.getLogger("Postgres")
        self.__connection = None

    def connect(self, config_file, environment):
        """Establishes a connection to a PostgreSQL database using the provided config file and environment."""
        try:
            config = configparser.ConfigParser()
            config.read(config_file)
            env_section = f'{environment}_postgres'
            self.__connection = psycopg.connect(
                host=config[env_section]['host'],
                port=config[env_section]['port'],
                user=config[env_section]['user'],
                password=config[env_section]['password'],
                dbname=config[env_section]['dbname'],
                row_factory=UniqueDictRowFactory
            )
            self.logger.debug(f"Connected to PostgreSQL database: {config[env_section]['dbname']}")

        except (psycopg.OperationalError, configparser.Error) as e:
            self.logger.error(f"Failed to connect to PostgreSQL: {str(e)}")
            raise

    def is_terminated(self):
        """Check if the PostgreSQL connection is closed."""
        return self.__connection.broken

    def close(self):
        """Closes the PostgreSQL database connection."""
        if self.__connection:
            self.__connection.close()
            self.logger.debug("PostgreSQL connection closed.")

    def commit(self):
        """It just commits!"""
        self.__connection.commit()

    def rollback(self):
        """Rolls back the transaction."""
        if self.__connection:
            self.__connection.rollback()

    def get_cursor(self, is_client_cursor = False):
        """Returns a generalized cursor object based on param for PostgreSQL."""
        try:
            if is_client_cursor:
                # Always return a client cursor (client-side)
                return PostgresCursor(self.__connection.cursor())
            else:
                return PostgresCursor(self.__connection.cursor(name=f"server_side_cursor_{uuid.uuid4()}", withhold = True))
        except Exception as e:
            self.logger.error(f"Failed to create PostgreSQL cursor: {str(e)}")
            raise

# Concrete class for PostgreSQL cursor
class PostgresCursor(GeneralCursor):
    def __init__(self, cursor):
        self.logger = logging.getLogger("Postgres")
        self.__cursor = cursor

    def __enter__(self):
        """Enters the context manager and returns the cursor."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Closes the cursor when exiting the context."""
        self.__cursor.close()

    def execute(self, query, params=None):
        """Executes a SQL query using the PostgreSQL cursor."""
        try:
            self.__cursor.execute(query, params)
            self.logger.debug("Executed query on PostgreSQL.")
        except psycopg.Error as e:
            self.logger.error(f"Failed to execute query on PostgreSQL: {str(e)}")
            raise

    def fetchone(self):
        """Fetches one row from the query result."""
        return self.__cursor.fetchone()

    def fetchall(self):
        """Fetches all rows from the query result."""
        return self.__cursor.fetchall()

    def fetchmany(self, size):
        """Fetches a specific number of rows from the query result."""
        return self.__cursor.fetchmany(size)

    @property
    def description(self):
        """Expose the description attribute from the underlying cursor."""
        return self.__cursor.description

# Concrete class for Oracle database connection
class OracleConnection(GeneralConnection):
    def __init__(self):
        self.logger = logging.getLogger("Oracle")
        self.__connection = None

    def connect(self, config_file, environment):
        """Establishes a connection to an Oracle database using the provided config file and environment."""
        try:
            config = configparser.ConfigParser()
            config.read(config_file)
            env_section = f'{environment}_oracle'
            if config[env_section].get('sid') is not None:
                dsn = oracledb.makedsn(
                    config[env_section]['host'],
                    config[env_section]['port'],
                    sid=config[env_section]['sid']
                )
            elif config[env_section].get('service_name') is not None:
                dsn = oracledb.makedsn(
                    config[env_section]['host'],
                    config[env_section]['port'],
                    service_name=config[env_section]['service_name']
                )
            else:
                dsn = None
            self.__connection = oracledb.connect(
                user=config[env_section]['user'],
                password=config[env_section]['password'],
                dsn=dsn
            )
            self.logger.debug(f"Connected to Oracle database: {config[env_section]}")

        except (oracledb.DatabaseError, configparser.Error) as e:
            self.logger.error(f"Failed to connect to Oracle: {str(e)}")
            raise

    def is_terminated(self):
        """Check if the Oracle connection is closed."""
        # Check the connection status by executing a simple query
        return not self.__connection.is_healthy()

    def close(self):
        """Closes the Oracle database connection."""
        if self.__connection:
            self.__connection.close()
            self.logger.debug("Oracle connection closed.")

    def get_cursor(self, is_client_cursor = False):
        """Returns a generalized cursor object for Oracle."""
        try:
            return OracleCursor(self.__connection.cursor())
        except Exception as e:
            self.logger.error(f"Failed to create Oracle cursor: {str(e)}")
            raise
    
    def commit(self):
        """It just commits!"""
        self.__connection.commit()

    def rollback(self):
        """Rolls back the transaction."""
        if self.__connection:
            self.__connection.rollback()
    

# Concrete class for Oracle cursor
class OracleCursor(GeneralCursor):
    def __init__(self, cursor):
        self.logger = logging.getLogger("Oracle")
        self.__cursor = cursor

    def __enter__(self):
        """Enters the context manager and returns the cursor."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Closes the cursor when exiting the context."""
        self.__cursor.close()

    def execute(self, query, params=None):
        """Executes a SQL query using the Oracle cursor."""
        try:
            self.__cursor.execute(query, params or {})
            self.logger.debug("Executed query on Oracle.")

        except oracledb.DatabaseError as e:
            self.logger.error(f"Failed to execute query on Oracle: {str(e)}")
            raise

    def fetchone(self):
        """Fetches one row from the query result."""
        self.__apply_row_factory()
        return self.__cursor.fetchone()

    def fetchall(self):
        """Fetches all rows from the query result."""
        self.__apply_row_factory()
        return self.__cursor.fetchall()

    def fetchmany(self, size):
        """Fetches a specific number of rows from the query result."""
        self.__apply_row_factory()
        return self.__cursor.fetchmany(size)

    @property
    def description(self):
        """Expose the description attribute from the underlying cursor."""
        return self.__cursor.description
    
    def __apply_row_factory(self):
        """Applies the row factory to format rows as dictionaries with column names as keys, handling duplicate column names."""
        # Extract column names from cursor description
        columns = [col[0] for col in self.__cursor.description]

        # Create a dictionary to count occurrences of column names
        field_count = {}
        unique_columns = []

        # Handle duplicate column names by appending a suffix
        for col_name in columns:
            if col_name in field_count:
                # If the column name already exists, append a unique identifier
                field_count[col_name] += 1
                unique_col_name = f"{col_name}_{field_count[col_name]}"
            else:
                field_count[col_name] = 0
                unique_col_name = col_name
            unique_columns.append(unique_col_name)

        # Set rowfactory to map unique column names to values
        self.__cursor.rowfactory = lambda *args: dict(zip(unique_columns, args))

# Transaction context manager
class Transaction:
    def __init__(self, connection):
        self.connection = connection
        self.success = False

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.connection.commit()
        else:
            self.connection.rollback()
            self.connection.logger.warning("Transaction rolled back due to an error.")

# SQLExecutor class manages the connection and execution of SQL queries
class SQLExecutor:
    def __init__(self, db_connection: GeneralConnection, config_file='config.ini', environment='test') -> None:
        """
        Initializes a new instance of SQLExecutor.
        This class handles database operations, including executing queries from files, saving results, and managing transactions.

        Parameters:
            db_connection (GeneralConnection): The database connection object used to interact with the database.
            config_file (str): The path to the configuration file for the database. Defaults to 'config.ini'.
            environment (str): The environment to use for the database connection (e.g., 'test', 'production'). Defaults to 'test'.

        Returns:
            None
        """
        self.logger = logging.getLogger("Executor")
        self.__db_connection = db_connection
        self.__config_file = config_file
        self.__environment = environment
        self.__db_connection.connect(config_file, environment)

    def __del__(self):
        """Closes the connection when the object is destroyed."""
        self.__db_connection.close()
    
    def connect(self, config_file, environment) -> None:
        """
        Establishes a connection to the database using the provided configuration file and environment.

        Parameters:
            config_file (str): The path to the configuration file for the database.
            environment (str): The environment to use for the database connection.

        Returns:
            None
        """
        self.__db_connection.connect(config_file, environment)

    def __get_cursor(self, is_client_cursor = None):
        """Returns a new cursor. Each method will call this to open a new cursor."""    
        if is_client_cursor:        
            return self.__db_connection.get_cursor(is_client_cursor)
        else:
            return self.__db_connection.get_cursor()
    
    def transaction(self) -> Transaction:
        """
        Returns a transaction context manager for managing database transactions.

        Returns:
            Transaction: A context manager for handling transactions.
        """
        return Transaction(self.__db_connection)

    def is_terminated(self) -> bool:
        """
        Checks if the database connection is terminated.

        Returns:
            bool: True if the connection is terminated, False otherwise.
        """
        return self.__db_connection.is_terminated()

    @retry(tries=3, delay=2, backoff=2, exceptions=(psycopg.OperationalError, oracledb.DatabaseError))
    def execute_file_and_save(self, file_name, result_file_path, result_file_type, batch_size=None, row_limit=None) -> None:
        """
        Executes SQL queries from a file and saves the results in the specified format.

        Parameters:
            file_name (str): The path to the file containing SQL queries.
            result_file_path (str): The base path where the results will be saved.
            result_file_type (FileType): The format to save the results (CSV, TXT, or Excel).
            batch_size (int, optional): The number of rows to fetch in each batch. Defaults to None.
            row_limit (int, optional): The maximum number of rows to fetch. Defaults to None.

        Returns:
            None
        """
        try:
            # Check if the connection is None or has been closed
            if self.__db_connection is None or self.__db_connection.is_terminated():
                self.logger.warning("Database connection is not active, attempting to reconnect.")
                self.__db_connection.connect(self.__config_file, self.__environment)
            else:

                with open(file_name, 'r', encoding="utf-8") as file:
                    queries = file.read().split(';')

                # This is to append count to the file name according to the number of queries a file have. Like (001, 002, 003, ...).
                preceding_zeros = len(str(len(queries)))

                # Loop through each query in the file
                for i,query in enumerate(queries):
                    query = query.strip()
                    if not query:  # Skip empty queries
                        continue
                    
                    # Check if the query includes pagination (e.g., /* PAGINATE SIZE <number> */)
                    is_paginate, query_page_size = SQLExecutor.extract_pagination_info(query)
                    is_rowlimit, query_row_limit = SQLExecutor.extract_row_limit_info(query)

                    # Apply limit based of query comments or parameter
                    apply_limit = query_row_limit if is_rowlimit else row_limit

                    # Apply batch_size based of query comments or parameter
                    apply_batch_size = query_page_size if is_paginate else batch_size

                    if (apply_batch_size and apply_batch_size <= 0) or (apply_limit and apply_limit <= 0):
                        self.logger.warning(f"No data has been fetched since batch size or row limit is zero or lesser for query number {i+1} and filename: {file_name}")
                        continue

                    # Handle queries with batch sizes (including 1)
                    if apply_batch_size:
                        # Fetch rows in specified batch sizes
                        batches = self.get_batches_by_query(query, apply_batch_size)
                        self.save_results_in_batches(batches, f"{result_file_path}_{str(i + 1).zfill(preceding_zeros)}", result_file_type, apply_limit=apply_limit, apply_batch_size=apply_batch_size)
                    else:
                        try:
                            with self.__get_cursor() as cursor:
                                cursor.execute(query)
                                # Fetch all rows at once if no batch size specified
                                if apply_limit:
                                    rows = cursor.fetchmany(apply_limit)
                                else:
                                    rows = cursor.fetchall()
                                if result_file_type == FileType.CSV:
                                    self.__save_to_csv(rows,f"{result_file_path}_{str(i+1).zfill(preceding_zeros)}", delimiter=',')    
                                elif result_file_type == FileType.TXT:
                                    self.__save_to_txt(rows,f"{result_file_path}_{str(i+1).zfill(preceding_zeros)}")    
                                elif result_file_type == FileType.EXCEL:
                                    self.__save_to_excel(rows,f"{result_file_path}_{str(i+1).zfill(preceding_zeros)}")
                        except Exception as e:
                            self.logger.error(f"Failed to execute and save results for query in file '{file_name}': {e}")
                            raise
                        finally:
                            self.__db_connection.commit()
                            
        except FileNotFoundError as e:
            self.logger.error(f"SQL file not found: {file_name}")
            raise
        except Exception as e:
            self.logger.error(f"Failed to execute SQL file: {str(e)}")
            raise

    @retry(tries=3, delay=2, backoff=2, exceptions=(psycopg.InterfaceError, oracledb.InterfaceError))
    def execute_folder_and_save(self,folder_path, result_save_path, result_file_type, batch_size=None, row_limit=None) -> None:
        """
        Executes SQL queries from all files in a folder and saves the results in the specified format.

        Parameters:
            folder_path (str): The path to the folder containing SQL files.
            result_save_path (str): The base path where the results will be saved.
            result_file_type (FileType): The format to save the results (CSV, TXT, or Excel).
            batch_size (int, optional): The number of rows to fetch in each batch. Defaults to None.
            row_limit (int, optional): The maximum number of rows to fetch. Defaults to None.

        Returns:
            None
        """

        # Check if the connection is None or has been closed
        if self.__db_connection is None or self.__db_connection.is_terminated():
            self.logger.warning("Database connection is not active, attempting to reconnect.")
            self.__db_connection.connect(self.__config_file, self.__environment)
        else:
            # Check if the provided path is a valid directory
            if not os.path.isdir(folder_path):
                self.logger.error(f"The provided path '{folder_path}' is not a valid directory.")
                return

            # Loop through all files in the folder
            for file_name in os.listdir(folder_path):
                file_path = os.path.join(folder_path, file_name)
                if os.path.isfile(file_path):
                    # Execute each file and save the results
                    self.execute_file_and_save(file_path, f"{result_save_path}{os.path.splitext(file_name)[0]}", result_file_type, batch_size, row_limit)

    def execute_query(self, query, params=None) -> None:
        """
        Executes a SQL query (INSERT, DELETE, UPDATE, etc.) with transaction management.

        Parameters:
            query (str): The SQL query to execute.
            params (dict, optional): Parameters for the SQL query. Defaults to None.

        Returns:
            None
        """
        with self.__get_cursor(is_client_cursor=True) as cursor:
            cursor.execute(query, params)

    def execute_file(self, file_name) -> None:
        """
        Executes SQL queries from a file, where each query is separated by a semicolon.

        Parameters:
            file_name (str): The path to the file containing SQL queries.

        Returns:
            None
        """
        try:
            with open(file_name, 'r', encoding="utf-8") as file:
                queries = file.read().split(';')
                queries = [query.strip() for query in queries if query.strip()]  # Clean empty queries
                for query in queries:
                    with self.__get_cursor(is_client_cursor=True) as cursor:
                        cursor.execute(query)

        except FileNotFoundError:
            self.logger.error(f"SQL file not found: {file_name}")
            raise
        except Exception as e:
            self.logger.error(f"Failed to execute SQL file: {str(e)}")
            raise

    @staticmethod
    def extract_pagination_info(query):
        """Extracts the pagination information from multiline comments using the pattern '/* PAGINATE SIZE <number> */'."""
        match = re.search(r'/\*\s*PAGINATE\s+SIZE\s+(\d+)\s*\*/', query, flags=re.IGNORECASE)
        if match:
            return True, int(match.group(1))  # Return pagination flag and page size as an integer
        return False, None  # Return False if no pagination information is found
    
    @staticmethod
    def extract_row_limit_info(query):
        """Extracts the row limit information from multiline comments using the pattern '/* ROW LIMIT <number> */'."""
        match = re.search(r'/\*\s*ROW\s+LIMIT\s+(\d+)\s*\*/', query, flags=re.IGNORECASE)
        if match:
            return True, int(match.group(1))  # Return pagination flag and page size as an integer
        return False, None  # Return False if no pagination information is found
    
    @staticmethod
    def extract_name_info(query):
        """Extracts the name information from multiline comments using the pattern '/* NAME <queryname> */'."""
        match = re.search(r'/\*\s*NAME\s*(\w+)\s*\*/', query, flags=re.IGNORECASE)
        if match:
            return True, match.group(1)  # Return pagination flag and page size as an integer
        return False, None  # Return False if no pagination information is found

    @retry(tries=3, delay=2, backoff=2, exceptions=(psycopg.OperationalError, oracledb.DatabaseError))
    def get_batches_by_query(self, query, page_size, params=None) -> list:
        """
        Fetches query results in batches.

        Parameters:
            query (str): The SQL query to execute.
            page_size (int): The number of rows to fetch per batch.
            params (optional): Parameters for the SQL query. Defaults to None.

        Returns:
            list: A list of rows fetched in batches.
        """
        try:
            # Check if the connection is None or has been closed
            if self.__db_connection is None or self.__db_connection.is_terminated():
                self.logger.warning("Database connection is not active, attempting to reconnect.")
                self.__db_connection.connect(self.__config_file, self.__environment)
            else:
                with self.__get_cursor() as cursor:
                    cursor.execute(query, params)
                    rows = cursor.fetchmany(page_size)
                    while rows:
                        yield rows
                        rows = cursor.fetchmany(page_size)
        except Exception as e:
            # Rollback if an error occurs
            self.logger.error(f"Error fetching batches: {e}")
            raise

        finally:
                self.__db_connection.commit()

    @retry(tries=3, delay=2, backoff=2, exceptions=(psycopg.OperationalError, oracledb.DatabaseError))
    def map_rows_to_objects(self, query, my_class, page_size, params=None) -> list:
        """
        Maps query results to instances of the specified class in batches.

        Parameters:
            query (str): The SQL query to execute.
            my_class (type): The class to map each row to.
            page_size (int): The number of rows to fetch per batch.
            params (dict, optional): Parameters for the SQL query. Defaults to None.

        Returns:
            list: A batches of list of class instances representing the query results.
        """
        try:
            if self.__db_connection is None or self.__db_connection.is_terminated():
                self.logger.warning("Database connection is not active, attempting to reconnect.")
                self.__db_connection.connect(self.__config_file, self.__environment)
            else:
                objects = []
                with self.__get_cursor() as cursor:
                    cursor.execute(query, params)
                    rows = cursor.fetchmany(page_size)
                    while rows:
                        for row in rows:
                            object = my_class()
                            for key, value in row.items():
                                # Set only the attributes that are present in the row dictionary
                                setattr(object, key, value)
                            objects.append(object)
                        yield objects
                        rows = cursor.fetchmany(page_size)
        except Exception as e:
            # Rollback if an error occurs
            self.logger.error(f"Error mapping objects by batches: {e}")
            raise

        finally:
            self.__db_connection.commit()

    def close(self) -> None:
        """
        Closes the database connection.

        Returns:
            None
        """
        self.__db_connection.close()

    @staticmethod
    def get_queries_from_file(file_name, index=None, Name=None) -> str or list:
        """
        Static Method Returns all queries or a specific query by index from a file.

        Parameters:
            file_name (str): The path to the file containing SQL queries.
            index (int, optional): The index of the query to return. If None, all queries are returned.

        Returns:
            str or list: A specific query as a string if index is provided, or a list of queries otherwise.
        """
        with open(file_name, 'r', encoding="utf-8") as file:
            queries = file.read().split(';')
            queries = [query.strip() for query in queries if query.strip()]

        if index is not None:
            # Return the query at the specified index if it exists
            if 0 <= index < len(queries):
                return queries[index]
            return None  # If the index is out of range, return None

        if Name is not None:
            # Return the query at the specified named query if it exists
            for query in queries:
                success, query_name = SQLExecutor.extract_name_info(query)
                if success and query_name == Name:
                    return query
            return None
        
        else:
            # Return all queries if no index is specified
            return queries
    
    def __ensure_directory_exists(self, file_path):
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
            self.logger.warning(f"No folder found, Creating {directory} folder.")
        
    def save_results(self, data, result_file, result_file_type: FileType, is_append=False, include_header=True, delimiter=',') -> None:
        """
        Saves data to a file in the specified format.

        Parameters:
            data (list): The data to save.
            result_file (str): The name of the file to save the data.
            result_file_type (FileType): The format to save the data (CSV, TXT, or Excel).
            is_append (bool): Whether to append to the file if it exists. Defaults to False.
            include_header (bool): Whether to include headers in the file. Defaults to True.

        Returns:
            None
        """
        self.__ensure_directory_exists(result_file)
        if result_file_type == FileType.CSV:
            self.__save_to_csv(data,result_file, delimiter, is_append, include_header)    
        elif result_file_type == FileType.TXT:
            self.__save_to_txt(data,result_file, is_append, include_header)  
        elif result_file_type == FileType.EXCEL:
            self.__save_to_excel(data,result_file, is_append, include_header)   

    def save_results_in_batches(self, batches, result_file, result_file_type: FileType, is_append=False, include_header=True, delimiter=',', apply_limit=None, apply_batch_size=None):
        """
        Saves batched data to a file efficiently, keeping the file open until all batches are processed.

        Parameters:
            batches (generator): A generator that yields batches of data.
            result_file (str): The name of the file to save the data.
            result_file_type (FileType): The format to save the data (CSV or Excel).
            include_header (bool): Whether to include headers in the file. Defaults to True.

        Returns:
            None
        """
        self.__ensure_directory_exists(result_file)
        if result_file_type == FileType.EXCEL:
            self.__save_batches_to_excel(batches, result_file, is_append, include_header, apply_limit, apply_batch_size)
        elif result_file_type == FileType.TXT:
            rows_fetched = 0
            batch_index = 0 
            for i,batch in enumerate(batches):
                # This logic will check that batch_size should be within the row limits.
                if apply_limit:
                    remaining_rows = apply_limit - rows_fetched
                    if remaining_rows < apply_batch_size:
                        batch = batch[:remaining_rows]
                self.__save_to_txt(batch,result_file, is_append= True if i!=0 else False, include_header=True if i==0 else False)
                rows_fetched += len(batch)
                batch_index += 1
                if apply_limit and rows_fetched >= apply_limit:
                    break  
        elif result_file_type == FileType.CSV:
            self.__save_batches_to_csv(batches, result_file, delimiter, is_append, include_header, apply_limit, apply_batch_size)

    def __save_to_csv(self, data, file_name, delimiter, is_append=False, include_header=True):
        """Saves data to a CSV file with column names."""
        try:
            if not data:  # Check if data is empty
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
        
            if isinstance(data, dict):  # If data is a single dictionary, wrap it in a list
                data = [data]

            # Check if the file name has a .csv extension, if not, add it
            if not file_name.lower().endswith('.csv'):
                file_name += '.csv'

            with open(file_name, 'a' if is_append else 'w', newline='',  encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=data[0].keys(), delimiter=delimiter)
                if include_header:
                    writer.writeheader()
                writer.writerows(data)
            self.logger.debug(f"Data saved to {file_name} as CSV.")
        except Exception as e:
            self.logger.error(f"Failed to save data to CSV: {str(e)}")
            raise


    def __save_batches_to_csv(self, all_batches, file_name, delimiter, is_append=False, include_header=True, apply_limit=None, apply_batch_size=None):
        """
        Saves all batches of data to a CSV file with column names.

        Parameters:
            all_batches (list): A list of batches where each batch is a list of dictionaries.
            file_name (str): The name of the CSV file.
            is_append (bool): Whether to append to the file if it exists.
            include_header (bool): Whether to include headers in the file.
            delimiter (str): The delimiter to use in the CSV file. Defaults to ','.

        Returns:
            None
        """
        try:
            if not all_batches:  # Check if there are any batches
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
            
            # Ensure the file name has a .csv extension
            if not file_name.lower().endswith('.csv'):
                file_name += '.csv'

            with open(file_name, 'a' if is_append else 'w', newline='', encoding='utf-8') as file:
                rows_fetched = 0
                batch_index = 0 
                for i, batch in enumerate(all_batches):
                    if not batch:  # Skip empty batches
                        continue
                    # This logic will check that batch_size should be within the row limits.
                    if apply_limit:
                        remaining_rows = apply_limit - rows_fetched
                        if remaining_rows < apply_batch_size:
                            batch = batch[:remaining_rows]
            
                    if isinstance(batch, dict):  # If data is a single dictionary, wrap it in a list
                        batch = [batch]

                    # Determine if the header should be included for this batch
                    write_header = include_header and not i > 0

                    # Write all rows to the CSV file
                    writer = csv.DictWriter(file, fieldnames=batch[0].keys(), delimiter=delimiter)
                    if write_header:
                        writer.writeheader()
                    writer.writerows(batch)

                    rows_fetched += len(batch)
                    batch_index += 1
                    if apply_limit and rows_fetched >= apply_limit:
                        break

            self.logger.debug(f"Data successfully saved to {file_name} as CSV.")
        except Exception as e:
            self.logger.error(f"Failed to save data to CSV: {str(e)}")
            raise

    def __save_to_txt(self, data, file_name, is_append=False, include_header=True):
        """Saves data to a TXT file with column names, using tab-separated values."""
        try:
            if not data:  # Check if data is empty
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
        
            if isinstance(data, dict):  # If data is a single dictionary, wrap it in a list
                data = [data]
            
            # Check if the file name has a .csv extension, if not, add it
            if not file_name.lower().endswith('.txt'):
                file_name += '.txt'

            with open(file_name, 'a' if is_append else 'w', newline='') as file:
                # Writing headers
                if include_header:
                    file.write('\t'.join(data[0].keys()) + '\n')
                # Writing data
                for row in data:
                    file.write('\t'.join(map(str, row.values())) + '\n')
            self.logger.debug(f"Data saved to {file_name} as TXT.")
        except Exception as e:
            self.logger.error(f"Failed to save data to TXT: {str(e)}")
            raise

    def __save_to_excel(self, data, file_name, is_append=False, include_header=True):
        """Saves data to an Excel file with column names."""
        try:
            if not data:  # Check if data is empty
                self.logger.warning(f"No data to save to {file_name}. The file will not be created.")
                return
        
            if not file_name.lower().endswith('.xlsx'):
                file_name += '.xlsx'

            if isinstance(data, dict):  # If data is a single dictionary, wrap it in a list
                data = [data]

            wb = None
            if is_append:
                try:
                    wb = load_workbook(file_name)
                except FileNotFoundError:
                    wb = Workbook()
            else:
                wb = Workbook()
            ws = wb.active
            # Writing headers
            if include_header:
                ws.append(list(data[0].keys()))
            # Writing data
            for row in data:
                ws.append(list(row.values()))
            wb.save(file_name)
            self.logger.debug(f"Data saved to {file_name} as Excel.")
        except Exception as e:
            self.logger.error(f"Failed to save data to Excel: {str(e)}")
            raise()

    def __save_batches_to_excel(self, batches, file_name, is_append=False, include_header=True, apply_limit=None, apply_batch_size=None):
        """
        Saves batched data to an Excel file using openpyxl, handling data in batches.

        Parameters:
            batches (generator): A generator that yields batches of data.
            file_name (str): The name of the Excel file.
            is_append (bool): Whether to append to an existing file. Defaults to False.
            include_header (bool): Whether to include headers in the file. Defaults to True.
            apply_limit (int): Maximum number of rows to write across all batches. Defaults to None.

        Returns:
            None
        """
        try:
            if not file_name.lower().endswith('.xlsx'):
                file_name += '.xlsx'

            # Load existing workbook or create a new one
            if is_append and os.path.exists(file_name):
                wb = load_workbook(file_name)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"

            ws = wb.active
            rows_fetched = 0 
            batch_index = 0 

            for batch in batches:
                if not batch:
                    continue

                if apply_limit:
                    remaining_rows = apply_limit - rows_fetched
                    if remaining_rows < apply_batch_size:
                        batch = batch[:remaining_rows]

                # Write headers if this is the first row of a new sheet
                if include_header and ws.max_row == 1:
                    ws.append(list(batch[0].keys()))

                # Write all rows in the batch at once
                for row in batch:
                    ws.append(list(row.values()))

                rows_fetched += len(batch)
                batch_index += 1
                if apply_limit and rows_fetched >= apply_limit:
                    break 

            # Save the workbook
            wb.save(file_name)
            self.logger.debug(f"Data successfully saved to {file_name} as Excel.")
        except Exception as e:
            self.logger.error(f"Failed to save batches to Excel: {str(e)}")
            raise

    